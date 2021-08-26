#!/usr/bin/env python3
"""
Author : hadev <hadev@localhost>
Date   : 2021-08-21
Purpose: Get Prices From Manufacturer
"""

import os
import argparse
from openpyxl.reader.excel import load_workbook


def get_workbook_to_search():
    files = dict(
        enumerate(
            [f for f in os.listdir(os.curdir) if os.path.isfile(f) and f[-4:] == "xlsx"]
        )
    )

    # list worksheets in same directory
    for key, file in files.items():
        print(str(key) + " " + file)

    # have user enter number key to workbook name than sheet to compare from
    print("Please enter number next to workbook")
    workbook_number = int(  # ask user for number
        input("Number left of workbook name: ")
    )
    print("------------------------------------\n")
    print(  # prints workbook name of selected number
        "Loading: " + files[workbook_number]
    )
    print("------------------------------------\n")
    workbook_object = load_workbook(  # loads workbook into variable
        files[workbook_number]
    )
    sheet_dict = dict(  # loads sheet names into a dict
        enumerate(workbook_object.sheetnames)
    )
    for (
        key,
        file,
    ) in sheet_dict.items():  # prints out index and sheet for user to select
        print(str(key) + " " + file)
        print("Please Enter Sheet we will work from")
        print("------------------------------------\n")
        sheet_number = int(input("number left of sheetname: "))
        print("------------------------------------\n")

    return (
        workbook_object,
        workbook_object[sheet_dict[sheet_number]],
        files[workbook_number],
    )


# --------------------------------------------------
def find_column(sheet_object, column_name):
    # ask if identifying column is known
    identifying_column = input(f"{column_name} Column?(leave blank if unknown): ")
    print("------------------------------------\n")

    if identifying_column == "":  # if nothing is entered
        looking_for_source_identifier = (  # we will proceed to look for identifying column
            True
        )
        column_to_look_in = 1  # start with row 1
    else:
        looking_for_source_identifier = (  # if something was entered we are NOT looking
            False
        )
        identifying_column = int(identifying_column)  # turn into integer

    # Looking for the identifying column if not known
    while looking_for_source_identifier:
        for cell in sheet_object[  # Going to list each rows value and row
            column_to_look_in
        ]:
            if cell.value is not None:
                print(str(cell.column) + ": " + str(cell.value))
            else:
                continue
            column_to_look_in += 1  # look in next row if entry is still blank

        # ask if identifying value is listed
        is_looking_for_source_identifier = input(
            "Enter number of Identifier(leave blank if not listed): "
        )
        print("------------------------------------\n")

        # if identifying column is entered it is assigned, if blank looking continues
        if is_looking_for_source_identifier != "":
            identifying_column = int(is_looking_for_source_identifier)
            looking_for_source_identifier = False

    return identifying_column


# --------------------------------------------------
def main():
    """Make a jazz noise here"""

    print("Lets Find Source Worksheet")
    print("------------------------------------")
    # unpack workbook assets from getting workbook, swo and swn to save, sso to work with
    (
        source_workbook_object,
        source_sheet_object,
        source_workbook_name,
    ) = get_workbook_to_search()
    sso = source_sheet_object
    source_identifying_column_number = find_column(  # get identifying column number
        sso, "Identifying"
    )
    source_list_price_column_number = find_column(  # get list price column number
        sso, "List Price"
    )
    sicn = source_identifying_column_number  # create smaller variables to make code shorter
    slpcn = source_list_price_column_number

    print("Lets Find Manufacturer Worksheet")
    print("------------------------------------")
    # unpacking manufacturer values, only need mpso rest are unpacked so there's no error
    (
        manufacturer_workbook_object,
        manufacturer_price_sheet_object,
        manufacturer_workbook_name,
    ) = get_workbook_to_search()
    mpo = manufacturer_price_sheet_object
    manufacturer_identifying_column_number = find_column(  # get id column number
        mpo, "Identifying"
    )
    manufacturer_list_price_column_number = find_column(  # get list price column
        mpo, "List Price"
    )
    # creating smaller variable for shorter code
    mlpcn = manufacturer_list_price_column_number
    micn = manufacturer_identifying_column_number

    # start iterations
    for source_identifying_column in sso.iter_cols(
        max_col=sicn, min_col=sicn  # Generates column from source sheet
    ):
        for (
            source_cell
        ) in source_identifying_column:  # for each cell in source sheet column
            if source_cell.value is not None:  # if cell is empty keep going
                # for each cell in the source sheet column we are going to go down manufacturer column to find a match
                for manufacturer_column in mpo.iter_cols(  # generates comparison column from manufacturer sheet
                    max_col=micn, min_col=micn
                ):
                    # for each cell in the column we are comparing to
                    for (  
                        manufacturer_cell
                    ) in (
                        manufacturer_column
                    ):
                        if (  # if the value from the source cell matches the current cell
                            source_cell.value == manufacturer_cell.value
                        ):
                            mcr = (  # take the row of matching manufacturer cell
                                manufacturer_cell.row
                            )
                            scr = (  # take the row of matching source cell
                                source_cell.row
                            )

                            # Get the value from the list price column of manufacturer sheet
                            manufacturer_list_price = mpo.cell(
                                row=mcr, column=mlpcn
                            ).value
                            # Prints the identifier cell that was matched, and price that was extracted
                            print(
                                str(source_cell.value)
                                + " price: "
                                + manufacturer_list_price
                            )
                            # set value of list price in source sheet list price column for row being searched
                            sso.cell(
                                row=scr, column=slpcn
                            ).value = manufacturer_list_price
                        else:
                            continue
            else:
                continue

        source_workbook_object.save(source_workbook_name)  # update the workbook
        print("Done Updating!")


# --------------------------------------------------
if __name__ == "__main__":
    main()
